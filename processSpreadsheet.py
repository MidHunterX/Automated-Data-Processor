import sqlite3      # SQLite DB operations
from openpyxl import Workbook       # Excel Workbook
from openpyxl.styles import Font    # Excel Font Styles
import function as fn
from function import var


def main():
    input_dir = var["input_dir"]
    db_file = var["db_file"]
    district_dataset = var["district_dataset"]
    spreadsheet_dir = fn.initNestedDir(input_dir, "Output Spreadsheet")

    for district in district_dataset:
        xlsx_file = f"{spreadsheet_dir}\\{district}.xlsx"
        generateOutputSpreadsheet(db_file, district, xlsx_file)
        print(f"✅ {xlsx_file} generated for {district}")


def generateOutputSpreadsheet(db_file, district, xlsx_file):
    """
    Arguments: (db_file, district, xlsx_filename)
        db_file: Path to SQLite Database
        district: Name of District as String
        xlsx_filename: Name of output .XLSX file eg: "Example.xlsx"

    Returns:
        Generated Excel.xlsx file
    """

    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    wb = Workbook()
    ws = wb.active

    # --------------------------------------------------------- [ SQL Queries ]

    if district != "Unknown":
        schoolSQL = """
        SELECT * FROM Schools WHERE District = ?
        """
        cursor.execute(schoolSQL, (district,))

    elif district == "Unknown":
        schoolSQL = """
        SELECT *
        FROM Schools
        WHERE District NOT IN (
        "Thiruvananthapuram", "Kollam", "Pathanamthitta", "Alappuzha",
        "Kottayam", "Idukki", "Ernakulam", "Thrissur", "Palakkad",
        "Malappuram", "Kozhikode", "Wayanad", "Kannur", "Kasargod" );
        """
        cursor.execute(schoolSQL)

    studentSQL = """
    SELECT * FROM Students WHERE SchoolID = ?
    """
    schools_table = cursor.fetchall()

    # -------------------------------------------------------- DISTRICT HEADING
    if district != "Unknown":
        ws.append([district.upper()])
    elif district == "Unknown":
        ws.append(["OTHER STATES"])
    row_number = ws.max_row
    for cell in ws[row_number]:
        cell.font = Font(bold=True, size=28)
    ws.append([""])

    # ------------------------------------------------------------ TABLE HEADER
    excel_header = ["Contacts", "Name", "Class", "AccNo", "IFSC", "Branch", "Amount"]
    ws.append(excel_header)
    row_number = ws.max_row
    for cell in ws[row_number]:
        cell.font = Font(bold=True, size=16)

    # ----------------------------------------------------- INSTITUTION DETAILS
    for school in schools_table:
        school_id = school[0]
        school_name = school[1]
        school_dist = school[2]
        school_place = school[3]
        school_no = school[4]
        school_mail = school[5]
        school_name_long = f"{school_name}, {school_place}"

        if district == "Unknown":
            school_name = f"{school_name} ({school_dist})"
            school_name_long = f"{school_name}, {school_place} ({school_dist})"

        if len(school_name_long) < 45:
            ws.append([school_name_long])
        else:
            ws.append([school_name])
        row_number = ws.max_row
        for cell in ws[row_number]:
            cell.font = Font(size=16)
        if school_no:
            ws.append([school_no])
        if school_mail:
            ws.append([school_mail])

        # ----------------------------------------------------- STUDENT DETAILS
        cursor.execute(studentSQL, (school_id,))
        student_table = cursor.fetchall()
        for student in student_table:
            st_name = student[2]
            st_cls = student[3]
            st_IFSC = student[4]
            st_acc = student[5]
            # st_holder = student[6]
            st_br = student[7]

            amt = fn.convertStdToAmount(st_cls)
            st_cls = fn.convertNumToStd(st_cls)

            excel_st_row = ["", st_name, st_cls, st_acc, st_IFSC, st_br, amt]
            ws.append(excel_st_row)

        # Space after each school entry
        ws.append([""])

    wb.save(xlsx_file)
    conn.close()
