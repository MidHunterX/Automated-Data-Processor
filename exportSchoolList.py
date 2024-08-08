import sqlite3      # SQLite DB operations
from openpyxl import Workbook       # Excel Workbook
from openpyxl.styles import Font    # Excel Font Styles
import function as fn
from function import var


def main():
    input_dir = var["input_dir"]
    db_file = var["db_file"]
    district_dataset = var["district_dataset"]
    spreadsheet_dir = fn.initNestedDir(input_dir, "Output Schools")

    for district in district_dataset:
        xlsx_file = f"{spreadsheet_dir}\\{district}.xlsx"
        generateOutputSpreadsheet(db_file, district, xlsx_file)
        print(f"✅ {xlsx_file} generated for {district}")


def generateOutputSpreadsheet(db_file, district, xlsx_file):
    """
    Arguments: (db_file, district, xlsx_filename)
        db_file: Path to SQLite Database
        district: Name of District as String
        xlsx_filename: Name of output .XLSX file eg: "Example.xlsx"

    Returns:
        Generated Excel.xlsx file
    """

    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    wb = Workbook()
    ws = wb.active

    # --------------------------------------------------------- [ SQL Queries ]

    if district != "Unknown":
        schoolSQL = """
        SELECT * FROM Schools WHERE District = ?
        """
        cursor.execute(schoolSQL, (district,))

    elif district == "Unknown":
        schoolSQL = """
        SELECT *
        FROM Schools
        WHERE District NOT IN (
        "Thiruvananthapuram", "Kollam", "Pathanamthitta", "Alappuzha",
        "Kottayam", "Idukki", "Ernakulam", "Thrissur", "Palakkad",
        "Malappuram", "Kozhikode", "Wayanad", "Kannur", "Kasargod" );
        """
        cursor.execute(schoolSQL)

    schools_table = cursor.fetchall()

    # -------------------------------------------------------- DISTRICT HEADING
    if district != "Unknown":
        ws.append([district.upper()])
    elif district == "Unknown":
        ws.append(["OTHER STATES"])
    row_number = ws.max_row
    for cell in ws[row_number]:
        cell.font = Font(bold=True, size=28)
    ws.append([""])

    # ------------------------------------------------------------ TABLE HEADER
    excel_header = ["Name", "Phone", "E-Mail"]
    ws.append(excel_header)
    row_number = ws.max_row
    for cell in ws[row_number]:
        cell.font = Font(bold=True, size=16)

    # ----------------------------------------------------- INSTITUTION DETAILS
    for school in schools_table:
        school_id = school[0]
        school_name = school[1]
        school_dist = school[2]
        school_place = school[3]
        school_no = school[4]
        school_mail = school[5]

        ws.append([school_name, school_no, school_mail])

    wb.save(xlsx_file)
    conn.close()
