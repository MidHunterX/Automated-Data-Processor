import sqlite3      # SQLite DB operations
from openpyxl import Workbook       # Excel Workbook
from openpyxl.styles import Font    # Excel Font Styles
import function as fn


def main(var):
    db_file = var["db_file"]
    xlsx_file = var["excel_file"]

    district = fn.getDistrictFromUser()
    generateOutputNEFT(db_file, district, xlsx_file)
    print(f"✅ {xlsx_file} generated for {district}")


def generateOutputNEFT(db_file, district, xlsx_file):
    """
    Arguments: (db_file, district, xlsx_filename)
        db_file: Path to SQLite Database
        district: Name of District as String
        xlsx_filename: Name of output .XLSX file eg: "Example.xlsx"

    Returns: xlsx file with bank NEFT format
    """

    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    wb = Workbook()
    ws = wb.active

    # --------------------------------------------------------- [ SQL Queries ]

    if district != "Unknown":
        schoolSQL = """
        SELECT * FROM Schools WHERE District = ?
        """
        cursor.execute(schoolSQL, (district,))

    elif district == "Unknown":
        schoolSQL = """
        SELECT *
        FROM Schools
        WHERE District NOT IN (
        "Thiruvananthapuram", "Kollam", "Pathanamthitta", "Alappuzha",
        "Kottayam", "Idukki", "Ernakulam", "Thrissur", "Palakkad",
        "Malappuram", "Kozhikode", "Wayanad", "Kannur", "Kasargod" );
        """
        cursor.execute(schoolSQL)

    studentSQL = """
    SELECT * FROM Students WHERE SchoolID = ?
    """
    schools_table = cursor.fetchall()

    # -------------------------------------------------------- DISTRICT HEADING
    if district != "Unknown":
        ws.append([district.upper()])
    elif district == "Unknown":
        ws.append(["OTHER STATES"])
    row_number = ws.max_row
    for cell in ws[row_number]:
        cell.font = Font(bold=True, size=28)
    ws.append([""])

    # ------------------------------------------------------------ TABLE HEADER
    excel_header = ["AccNo", "AccType", "AccTitle", "Addr", "IFSC", "Amount"]
    ws.append(excel_header)
    row_number = ws.max_row
    for cell in ws[row_number]:
        cell.font = Font(bold=True, size=16)

    # --------------------------------------------------------- STUDENT DETAILS
    for school in schools_table:

        school_id = school[0]
        cursor.execute(studentSQL, (school_id,))
        student_table = cursor.fetchall()

        for student in student_table:
            # st_name = student[2]
            st_class = student[3]
            st_IFSC = student[4]
            st_accno = student[5]
            st_holder = student[6]
            st_branch = student[7]

            amount = fn.convertStdToAmount(st_class)

            st_row = [st_accno, 10, st_holder, st_branch, st_IFSC, amount]
            ws.append(st_row)

    wb.save(xlsx_file)
    conn.close()
